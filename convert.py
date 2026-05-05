"""
Converts Data/prices.xlsx to Data/prices.json for the web frontend.
"""

import json
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path("Data/prices.xlsx")
JSON_PATH = Path("Data/prices.json")


def main():
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        timestamp, energy_price, feed_in_rate = row[0], row[1], row[2]
        rows.append(
            {
                "timestamp": timestamp,
                "energy_price_c_kwh": energy_price,
                "feed_in_rate_c_kwh": feed_in_rate,
            }
        )

    output = {"postcode": "2600", "data": rows}

    JSON_PATH.write_text(json.dumps(output, indent=2))
    print(f"Wrote {len(rows)} rows to {JSON_PATH}")


if __name__ == "__main__":
    main()
