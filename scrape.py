"""
Scrapes https://www.amber.com.au/ for live energy prices in postcode 2600
and appends a timestamped row to a local Excel file (committed back to the repo).
"""

import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

AMBER_URL = "https://www.amber.com.au/"
POSTCODE = "2600"
XLSX_PATH = Path("Data/prices.xlsx")

# Matches values like "12.3 c/kWh", "-5.1¢/kWh", "12.3 cents/kWh"
PRICE_PATTERN = re.compile(
    r"([\-\d]+\.?\d*)\s*(?:c|¢|cents?)\s*/\s*kWh", re.IGNORECASE
)


def _extract_prices_from_text(text: str) -> tuple[float | None, float | None]:
    """Return (energy_price, feed_in_rate) parsed from a block of page text."""
    lines = text.split("\n")
    energy_price: float | None = None
    feed_in_rate: float | None = None

    for i, line in enumerate(lines):
        line_lower = line.lower()
        context = "\n".join(lines[max(0, i - 2) : i + 6])

        # Energy / usage price
        if energy_price is None and any(
            kw in line_lower
            for kw in ["energy price", "general usage", "usage price", "live price"]
        ):
            m = PRICE_PATTERN.search(context)
            if m:
                energy_price = float(m.group(1))
                print(f"  Found energy price: {energy_price} c/kWh")

        # Feed-in rate
        if feed_in_rate is None and any(
            kw in line_lower
            for kw in ["feed-in", "feedin", "feed in", "solar", "export"]
        ):
            m = PRICE_PATTERN.search(context)
            if m:
                feed_in_rate = float(m.group(1))
                print(f"  Found feed-in rate: {feed_in_rate} c/kWh")

    # Fallback: collect all price values in document order
    if energy_price is None or feed_in_rate is None:
        all_vals = [float(m.group(1)) for m in PRICE_PATTERN.finditer(text)]
        unique_vals = list(dict.fromkeys(all_vals))
        if energy_price is None and unique_vals:
            energy_price = unique_vals[0]
            print(f"  Fallback energy price: {energy_price} c/kWh")
        if feed_in_rate is None and len(unique_vals) >= 2:
            feed_in_rate = unique_vals[1]
            print(f"  Fallback feed-in rate: {feed_in_rate} c/kWh")

    return energy_price, feed_in_rate


def scrape_prices() -> dict:
    """Navigate amber.com.au, enter postcode, and return live energy prices."""
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/136.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        page = context.new_page()

        print(f"  Navigating to {AMBER_URL} ...")
        page.goto(AMBER_URL, timeout=60_000, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=15_000)
        except Exception:
            print("  networkidle wait timed out; continuing anyway")
        page.wait_for_timeout(2_000)  # brief settle for client-side rendering

        # Find the postcode input
        postcode_input = None
        selectors = [
            'input[placeholder*="postcode" i]',
            'input[placeholder*="post code" i]',
            'input[placeholder*="suburb" i]',
            'input[name*="postcode" i]',
            'input[id*="postcode" i]',
            'input[aria-label*="postcode" i]',
            'input[type="number"][maxlength="4"]',
            'input[type="text"][maxlength="4"]',
            'input[type="text"]',
            'input[type="number"]',
        ]
        for sel in selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=2_000):
                    postcode_input = el
                    print(f"  Found postcode input via selector: {sel}")
                    break
            except Exception:
                continue

        if postcode_input is None:
            # Save screenshot for debugging
            page.screenshot(path="debug_screenshot.png")
            raise RuntimeError(
                "Could not find a postcode input on amber.com.au. "
                "A debug screenshot has been saved to debug_screenshot.png"
            )

        # Enter postcode and submit
        postcode_input.click()
        postcode_input.fill(POSTCODE)
        postcode_input.press("Enter")

        # Allow page to update after postcode entry
        print("  Waiting for price data to load...")
        try:
            page.wait_for_load_state("networkidle", timeout=30_000)
        except Exception:
            pass  # Continue even if networkidle times out
        page.wait_for_timeout(2_000)  # Extra settle time for React rendering

        # Also try clicking any visible "Submit" / "See prices" / "Check" button
        for btn_text in ["see prices", "check", "get prices", "submit", "go"]:
            try:
                btn = page.get_by_role("button", name=re.compile(btn_text, re.IGNORECASE))
                if btn.first.is_visible(timeout=1_000):
                    btn.first.click()
                    page.wait_for_load_state("networkidle", timeout=20_000)
                    break
            except Exception:
                continue

        page_text = page.inner_text("body")
        energy_price, feed_in_rate = _extract_prices_from_text(page_text)

        if energy_price is None and feed_in_rate is None:
            page.screenshot(path="debug_screenshot.png")

        browser.close()

    if energy_price is None and feed_in_rate is None:
        raise RuntimeError(
            "Could not extract any price data from amber.com.au. "
            "The page structure may have changed. "
            "Check debug_screenshot.png if available."
        )

    return {
        "energy_price_c_kwh": energy_price,
        "feed_in_rate_c_kwh": feed_in_rate,
    }


def main():
    print(f"Scraping Amber energy prices for postcode {POSTCODE} ...")
    data = scrape_prices()
    print(
        f"  Live energy price : {data['energy_price_c_kwh']} c/kWh\n"
        f"  Live feed-in rate : {data['feed_in_rate_c_kwh']} c/kWh"
    )

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        print("  No existing file — creating new spreadsheet.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Amber Prices"
        ws.append(["Timestamp", "Energy Price (c/kWh)", "Feed-In Rate (c/kWh)"])

    ws.append([timestamp, data["energy_price_c_kwh"], data["feed_in_rate_c_kwh"]])

    wb.save(XLSX_PATH)
    print(f"Done! Row added at {timestamp} → {XLSX_PATH}")


if __name__ == "__main__":
    main()
